from tkinter import *
from PIL import Image, ImageTk
import tkinter.messagebox as tmsg
import os
from openpyxl import *
from scripts import update as up
from scripts import detect as dt
from scripts import Dates as D
import CRUD as op
import datetime



def clear():  # clear the content of text entry box 
    name.delete(0, END) 
    enroll.delete(0, END) 
    course.delete(0, END) 
    semester.delete(0, END) 
    section.set('') 
    contact.delete(0, END) 
    email.delete(0, END) 

def update_in_attendace_sheet(semester,section,name,contact,email,course,enroll):
    if semester.get() == '1' or semester.get() == '2':
        year = 'first_year'
    elif semester.get() == '3' or semester.get() == '4':
        year = 'second_year'
    elif semester.get() == '5' or semester.get() == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'
    
    wb = load_workbook(f'../data/excel/{year}/{year}_{semester.get()}sem_IT{section.get()}1.xlsx') 
    sheet = wb.active 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 25

    # write given data to an excel spreadsheet 
    
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Enrollment"
    
    
    
    current_row = sheet.max_row 
    current_column = sheet.max_column 

    sheet.cell(row=current_row + 1, column=1).value = name.get() 
    sheet.cell(row=current_row + 1, column=2).value = enroll.get() 
 
     
    wb.save(f'../data/excel/{year}/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    

def update_in_data_sheet(semester,section,name,contact,email,course,enroll):
    if semester.get()=='':
        return

    # if semester.get() == '1' or semester.get() == '2':
    #     year = 'first_year'
    # elif semester.get() == '3' or semester.get() == '4':
    #     year = 'second_year'
    # elif semester.get() == '5' or semester.get() == '6':
    #     year = 'third_year'
    # else:
    #     # year = 'fourth_year'
    #     return
    print(semester.get())
    #wb = load_workbook(f'../data/excel/{year}/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    wb = load_workbook(f'data/excel/first_year/first_year_1sem_IT1.xlsx') 
    sheet = wb.active 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 25

    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Enrollment"
    sheet.cell(row=1, column=3).value = "Course"
    sheet.cell(row=1, column=4).value = "Section"
    sheet.cell(row=1, column=5).value = "Semester"
    sheet.cell(row=1, column=6).value = "Contact No"
    sheet.cell(row=1, column=7).value = "Email id"
    
    
    current_row = sheet.max_row 
    current_column = sheet.max_column 

    sheet.cell(row=current_row + 1, column=1).value = name.get()
    sheet.cell(row=current_row + 1, column=2).value = enroll.get()
    sheet.cell(row=current_row + 1, column=3).value = course.get()
    sheet.cell(row=current_row + 1, column=4).value = section.get()
    sheet.cell(row=current_row + 1, column=5).value = semester.get()
    sheet.cell(row=current_row + 1, column=6).value = contact.get()
    sheet.cell(row=current_row + 1, column=7).value = email.get()
     
    #wb.save(f'../data/excel/{year}/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    wb.save(f'data/excel/first_year/first_year_1sem_IT1.xlsx') 
    # clear()

    print("form submitted")