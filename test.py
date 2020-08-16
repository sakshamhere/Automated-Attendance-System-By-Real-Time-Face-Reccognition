# import libraries
################################### lib for TKINTER GUI
from tkinter import *
import tkinter.messagebox as tmsg
################################### importing scripts
from scripts import update as up
from scripts import detect as dt
from scripts import Dates as D
################################### lib to perform CRUD operations in csv/excel
from PIL import Image, ImageTk
from openpyxl import *
import pandas as pd
import numpy as np
import csv
################################### other dependencies
import datetime
import time
import os

'''*************************************************************** Code for CRUD operations starts ****************************************************************''' 

def clear():  # clear the content of text entry box 
    try:
        name.delete(0, END) 
        enroll.delete(0, END) 
        course.delete(0, END) 
        semester.delete(0, END) 
        section.delete(0, END) 
        contact.delete(0, END) 
        email.delete(0, END) 
    except:
        pass

def modify_in_record(sem,sec,new):
    if sem == '1' or sem == '2':
        year = 'first_year'
    elif sem == '3' or sem == '4':
        year = 'second_year'
    elif sem == '5' or sem == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'
    # print(new,sem,sec)
    filename = f'data/excel/{year}/{year}_{sem}sem_IT{sec}.xlsx'
    df = pd.read_excel(filename,index=False)
    df.to_csv('./data.csv')
    newlines = []
    with open('data.csv','r') as f:
        data = csv.reader(f)
        lines = list(data)
        for line in lines:
            if line[1]==new[3]:
                print('Initial record',line)
                line[2],line[3],line[4]= new[0],new[5],new[4]
                newlines.append(line[1:])
            else:

                newlines.append(line[1:])

    with open('current.csv','w') as g:
        writer = csv.writer(g,lineterminator='\n')
        writer.writerows(newlines)

    df = pd.read_csv('current.csv')
    f = pd.ExcelWriter(filename)
    df.to_excel(f,index=False)
    f.save()
    print('Record updated for',new[0])
    
def Register_in_record():

    if semester.get() == '1' or semester.get() == '2':
        year = 'first_year'
    elif semester.get() == '3' or semester.get() == '4':
        year = 'second_year'
    elif semester.get() == '5' or semester.get() == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'

    print(sectionn.get())

    wb = load_workbook(f'data/excel/{year}/{year}_{semester.get()}sem_IT{sectionn.get()}.xlsx') 

    sheet = wb.active 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10

    sheet.cell(row=1, column=1).value = "Enrollment Number"
    sheet.cell(row=1, column=2).value = "Name"
    sheet.cell(row=1, column=3).value = "Email ID"
    sheet.cell(row=1, column=4).value = "WhatsApp Mobile Number"
    
    current_row = sheet.max_row 
    current_column = sheet.max_column 

    sheet.cell(row=current_row + 1, column=1).value = enroll.get()
    sheet.cell(row=current_row + 1, column=2).value = name.get()
    sheet.cell(row=current_row + 1, column=3).value = email.get() 
    sheet.cell(row=current_row + 1, column=4).value = contact.get() 
    
    wb.save(f'data/excel/{year}/{year}_{semester.get()}sem_IT{sectionn.get()}.xlsx') 

    print(name.get()," is Registered successfully")

    clear() # clearing field

def Retrive(roll,sem,year,section):
    if sem == '1' or sem == '2':
        year = 'first_year'
    elif sem == '3' or sem == '4':
        year = 'second_year'
    elif sem == '5' or sem == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'

    filename = f'data/excel/{year}/{year}_{sem}sem_IT{section}.xlsx'
    df = pd.read_excel(filename,index=False)
    l = df[df['Enrollment Number']==roll].values.tolist()

    return l

def delete(enroll,sem,sec):
    if sem == '1' or sem == '2':
        year = 'first_year'
    elif sem == '3' or sem == '4':
        year = 'second_year'
    elif sem == '5' or sem == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'
    # print(enroll)
    filename = f'data/excel/{year}/{year}_{sem}sem_IT{sec}.xlsx'
    df = pd.read_excel(filename,index=False)
    df.to_csv('./data.csv')
    nlines = []
    with open('data.csv','r') as f:
        data = csv.reader(f)
        lines = list(data)
        for line in lines:
            # print(line)
            if enroll in line:
                print('Deleting record of ',line[2])
                pass
            else:
                nlines.append(line[1:])
    
    with open('current.csv','w') as g:
        writer = csv.writer(g,lineterminator='\n')
        writer.writerows(nlines)
    time.sleep(3)

    

    df = pd.read_csv('current.csv')
    f = pd.ExcelWriter(filename)
    df.to_excel(f,index=False)
    f.save()
    print('Record deleted successfully')

def updatef():
    filename = 'data/Attendance_xlsx/third_year_5sem_IT2.xlsx'
    with open('data.csv') as f:
        data = csv.reader(f)
        lines = list(data)
        if lines[1][0] == '0':
            
            for line in lines:
                line.pop(0)
            with open('data.csv','w') as g:
                writer = csv.writer(g,lineterminator='\n')
                writer.writerows(lines)
            
            df = pd.read_csv('data.csv')
            df.to_excel(filename,index = False)
            mail()

        else:
            print('Already Updated')

'''*************************************************************** Code for CRUD operations Ends ****************************************************************''' 
'''*************************************************************** Code for GUI starts ****************************************************************''' 

'''*************************************************************** initialize Window ****************************************************************''' 

# creating tkinter window
root = Tk()

# creating fixed geometry of the 
# tkinter window with dimensions 150x200 
root.geometry("150x200")
root.maxsize(580,550)
root.minsize(580,550)

#providing a title
root.title("Auto Attendance..")

image = Image.open("GUI/register.jpg")
image=image.resize((580,600), Image.ANTIALIAS)
photo = ImageTk.PhotoImage(image)

''' ************************************************************* Code for Login window starts ********************************************************'''
global login_by
Admin_id="admin"
Admin_pass="admin@123"

def Login():
    global login_by
    lid1=lid.get()
    lpass1=lpass.get()
    print(lid.get(),lpass.get())
    if(lid1==Admin_id and lpass1==Admin_pass):
        login_by="Admin"
        Proceed_menu()
    else:
        login_by=""
        Proceed_menu()

l_login=Label(image=photo)
f_login=Frame(l_login,pady="25",padx="25") #cretaing a Frame which can expand according to the size of the window

lb0 =Label(f_login,text="Enter Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
lb1 =Label(f_login,text="Enter ID: ",font="lucida 10 bold").grid(column=0,row=2,pady="4")

lid =StringVar()
e1 =Entry(f_login,textvariable=lid,width="28").grid(column=1,row=2)
lb2 =Label(f_login,text="Enter Password: ",font="lucida 10 bold").grid(column=0,row=3,pady="4")

lpass=StringVar()
e2=Entry(f_login,textvariable=lpass,width="28").grid(column=1,row=3)
btn=Button(f_login,text="login",bg="green",fg="white",width="10",font="lucida 10 bold",command=Login)
btn.grid(columnspan=3,row=5,pady="10")

f_login.pack(pady="165")

l_login.pack(ipadx="100",fill=BOTH)

''' ************************************************************* Code for Login window Ends ********************************************************'''

''' ************************************************************* Code for Menu Widgit starts ********************************************************'''


def student(x):
    l2.pack_forget()
    l1.pack_forget()
    l.pack(ipadx="100",fill=BOTH)
    if(x==1):
        f4.pack_forget()
        f3.pack_forget()
        f2.pack_forget()
        f21.pack_forget()
        f31.pack_forget()
        f1.pack(pady="110")
    if(x==2):
        f1.pack_forget()
        f3.pack_forget()
        f4.pack_forget()
        f31.pack_forget()
        f21.pack_forget()
        f2.pack(pady="115")
    
    if(x==3):
        f1.pack_forget()
        f2.pack_forget()
        f21.pack_forget()
        f4.pack_forget()
        f31.pack_forget()
        f3.pack(pady="115")
    if(x==4):
        f4.pack(pady="115")
        f1.pack_forget()
        f3.pack_forget()
        f31.pack_forget()
        f2.pack_forget()
        f21.pack_forget()

def attendance(y):
    l.pack_forget()
    l2.pack_forget()
    l1.pack(ipadx="150",fill=BOTH)
    if(y==1):
        fa.pack_forget()
        f98.pack(pady="120")
    if(y==2):
        f98.pack_forget()
        fa.pack(pady="135")

        
def more(z):
    l.pack_forget()
    l1.pack_forget()
    l2.pack(ipadx="100",fill=BOTH)
    
def admin(x):
    if(x==1):
        print("i will train")
    elif(x==2):
        print("i will open excel of teachers")

  
def Proceed_menu():
    global login_by
    l_login.pack_forget()
    
    mainmenu = Menu(root)
    
    m1 = Menu(mainmenu, tearoff=0)
    m1.add_command(label="Register", command=lambda: student(1))
    m1.add_command(label="View", command=lambda: student(2))
    m1.add_separator()
    m1.add_command(label="Update", command=lambda: student(3))
    m1.add_command(label="Delete", command=lambda: student(4))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="Student", menu=m1)
    
    m2 = Menu(mainmenu, tearoff=0)
    m2.add_command(label="Detect", command=lambda:attendance(1))
    m2.add_separator()
    m2.add_command(label="View Excel", command=lambda:attendance(2))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="Attendance", menu=m2)
   
    if(login_by=="Admin"):
        m4 = Menu(mainmenu, tearoff=0)
        m4.add_command(label="Train",command=lambda:admin(1))
        m4.add_command(label="Teacher Excel",command=lambda:admin(2))
        root.config(menu=mainmenu)
        mainmenu.add_cascade(label="Admin", menu=m4)
    
    m3 = Menu(mainmenu, tearoff=0)
    m3.add_command(label="Help", command=lambda:more(1))
    m3.add_command(label="About Us", command=lambda:more(1))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="More", menu=m3)
    
''' ************************************************************* Code for Menu Widgit Ends ********************************************************'''

''' ************************************************************* Code for Registeration starts ********************************************************'''


l = Label(image=photo)
f1=Frame(l,pady="5",padx="25")
l0=Label(f1,text="Registration Form",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
l1=Label(f1,text="Name : ",font="lucida 10 bold").grid(column=0,row=1,pady="4")
name=Entry(f1,width="28")
name.grid(column=1,row=1)
l2=Label(f1,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
enroll=Entry(f1,width="28")
enroll.grid(column=1,row=2)
l3=Label(f1,text="Course : ",font="lucida 10 bold").grid(column=0,row=3,pady="4")
course=Entry(f1,width="28")
course.grid(column=1,row=3)

l32=Label(f1,text="Section : ",font="lucida 10 bold").grid(column=0,row=4,pady="4")

sectionn=Entry(f1,width="28")
sectionn.grid(column=1,row=4)


l33=Label(f1,text="Sem : ",font="lucida 10 bold").grid(column=0,row=5,pady="4")
# inte = IntVar()
semester=Entry(f1,width="28")#textvariable=inte
semester.grid(column=1,row=5)

l5=Label(f1,text="Contact No : ",font="lucida 10 bold").grid(column=0,row=6,pady="4")
contact=Entry(f1,width="28")
contact.grid(column=1,row=6)

l6=Label(f1,text="Email : ",font="lucida 10 bold").grid(column=0,row=7,pady="4")
email=Entry(f1,width="28")
email.grid(column=1,row=7)

btn=Button(f1,text="Submit",bg="green",fg="white",width="10",font="lucida 10 bold",command = Register_in_record)
btn.grid(columnspan=3,row=8,pady="10")
f1.pack(pady="110")

''' ************************************************************* Code for Registeration ENDS ********************************************************'''
    
''' ************************************************************* Code for view details starts ********************************************************'''

global Nv,Mv,Env,Nv
Ev,Mv,Env,Nv =0,0,0,0
def view():

    l = Retrive(venroll.get(),vsem.get()[0],vyear.get()[0],vsection.get()[0])
    print('Record',l)
    Ev,Nv,Env,Mv = l[0][0],l[0][1],l[0][2],l[0][3]
   
    l0=Label(f21,text="Student Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")   
    l1=Label(f21,text="Name : ",font="lucida 10 bold").grid(column=0,row=1,pady="4")
    l11=Label(f21,text=Nv,width="28").grid(column=1,row=1)
    l2=Label(f21,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
    l22=Label(f21,text=Ev,width="28").grid(column=1,row=2)
    l3=Label(f21,text="SEM : ",font="lucida 10 bold").grid(column=0,row=3,pady="4")
    l33=Label(f21,text=vsem.get()[0],width="28").grid(column=1,row=3)
    l4=Label(f21,text="YEAR : ",font="lucida 10 bold").grid(column=0,row=4,pady="4")
    l44=Label(f21,text=vyear.get()[0],width="28").grid(column=1,row=4)
    l5=Label(f21,text="Contact No : ",font="lucida 10 bold").grid(column=0,row=5,pady="4")
    l55=Label(f21,text=Mv,width="28").grid(column=1,row=5)
    l6=Label(f21,text="Email : ",font="lucida 10 bold").grid(column=0,row=6,pady="4")
    l66=Label(f21,text=Env,width="28").grid(column=1,row=6)

    f2.pack_forget()
    f21.pack(pady="100")

def back():
    f21.pack_forget()
    f2.pack(pady="115")

f2=Frame(l,pady="25",padx="25")
l0=Label(f2,text="Student details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
l2=Label(f2,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
venroll = StringVar()
e2=Entry(f2,textvariable=venroll,width="28").grid(column=1,row=2)

l3=Label(f2,text="Year",font="lucida 10 bold").grid(column=0,row=3,pady="4")
vyear = StringVar()
vyear.set("1st Year") # default value
w = OptionMenu(f2,vyear, "1st Year", "2nd Year", "3rd Year","4th Year").grid(column=1,row=3,pady="4")

l4=Label(f2,text="Sem",font="lucida 10 bold").grid(column=0,row=4,pady="4")
vsem = StringVar()
vsem.set("1st_sem") # default value
w1 = OptionMenu(f2,vsem,"1st_sem","2nd_sem","3rd_sem","4th_sem","5th_sem","6th_sem","7th_sem","8th_sem").grid(column=1,row=4,pady="4")

l5=Label(f2,text="Section",font="lucida 10 bold").grid(column=0,row=5,pady="4")
vsection = StringVar()
# vsection.set("IT-01") # default value
w2 = OptionMenu(f2,vsection, "1", "2").grid(column=1,row=5,pady="4")

btn=Button(f2,text="OK",bg="green",fg="white",width="10",font="lucida 10 bold",command=view)
btn.grid(columnspan=3,row=7,pady="20")
f2.pack(pady="115")



f21=Frame(l,pady="25",padx="25")   
l0=Label(f21,text="Student Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")   
l1=Label(f21,text="Name : ",font="lucida 10 bold").grid(column=0,row=1,pady="4")
l11=Label(f21,text=Nv,width="28").grid(column=1,row=1)
l2=Label(f21,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
l22=Label(f21,text=Ev,width="28").grid(column=1,row=2)
l3=Label(f21,text="SEM : ",font="lucida 10 bold").grid(column=0,row=3,pady="4")
l33=Label(f21,text=vsem.get()[0],width="28").grid(column=1,row=3)
l4=Label(f21,text="YEAR : ",font="lucida 10 bold").grid(column=0,row=4,pady="4")
l44=Label(f21,text=vyear.get()[0],width="28").grid(column=1,row=4)
l5=Label(f21,text="Contact No : ",font="lucida 10 bold").grid(column=0,row=5,pady="4")
l55=Label(f21,text=Mv,width="28").grid(column=1,row=5)
l6=Label(f21,text="Email : ",font="lucida 10 bold").grid(column=0,row=6,pady="4")
l66=Label(f21,text=Env,width="28").grid(column=1,row=6)
btn=Button(f21,text="Back",bg="green",fg="white",width="10",font="lucida 10 bold",command=back)
btn.grid(columnspan=3,row=7,pady="20")
# f21.pack(pady="100")   



''' ************************************************************* Code for view details Ends ********************************************************'''

''' ************************************************************* Code for  Update starts ********************************************************'''


def update():
    
    f3.pack_forget()
    f31.pack(pady="100")    
    
def cancel():
    f31.pack_forget()
    f3.pack(pady="115")
def update_details():
    new = [uname.get(),usem.get(),usec.get(),uenrollment.get(),ucontact.get(),uemail.get()]
    modify_in_record(seme.get()[0],section.get(),new)
    

f31=Frame(l,pady="25",padx="25")
l0=Label(f31,text="Update Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15") 
l1=Label(f31,text="Name : ",font="lucida 10 bold").grid(column=0,row=1,pady="4")

uname=StringVar()
e1=Entry(f31,textvariable=uname,width="28").grid(column=1,row=1) 
l2=Label(f31,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")

uenrollment=StringVar()
e2=Entry(f31,textvariable=uenrollment,width="28").grid(column=1,row=2) 
l3=Label(f31,text="Sem : ",font="lucida 10 bold").grid(column=0,row=3,pady="4")

usem = StringVar()   
e3=Entry(f31,textvariable=usem,width="28").grid(column=1,row=3) 
l4=Label(f31,text="section : ",font="lucida 10 bold").grid(column=0,row=4,pady="4")

usec=StringVar()
e4=Entry(f31,textvariable=usec,width="28").grid(column=1,row=4) 
l5=Label(f31,text="Contact No : ",font="lucida 10 bold").grid(column=0,row=5,pady="4")

ucontact=StringVar()
e5=Entry(f31,textvariable=ucontact,width="28").grid(column=1,row=5) 
l6=Label(f31,text="Email : ",font="lucida 10 bold").grid(column=0,row=6,pady="4")

uemail=StringVar()
e6=Entry(f31,textvariable=uemail,width="28").grid(column=1,row=6) 
btn=Button(f31,text="Submit",bg="green",fg="white",width="10",font="lucida 10 bold",command=update_details)
btn.grid(column=0,row=7,pady="20")    
btn1=Button(f31,text="cancel",bg="green",fg="white",width="10",font="lucida 10 bold",command=cancel)
btn1.grid(column=1,row=7,pady="20")    
f31.pack(pady="100")


f3=Frame(l,pady="25",padx="25")
l0=Label(f3,text="Update details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
l2=Label(f3,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
en = StringVar()
e2=Entry(f3,textvariable=en,width="28").grid(column=1,row=2)

l3=Label(f3,text="Year",font="lucida 10 bold").grid(column=0,row=3,pady="4")
year = StringVar()
# year.set("1st Year") # default value
w = OptionMenu(f3,year, "1st Year", "2nd Year", "3rd Year","4th Year").grid(column=1,row=3,pady="4")

l4=Label(f3,text="Sem",font="lucida 10 bold").grid(column=0,row=4,pady="4")
seme = StringVar()
# sem.set("1st sem") # default value
w1 = OptionMenu(f3,seme,"1st sem","2nd sem","3rd sem","4th sem","5th sem","6th sem","7th sem","8th sem").grid(column=1,row=4,pady="4")

l5=Label(f3,text="Section",font="lucida 10 bold").grid(column=0,row=5,pady="4")
section = StringVar()
# section.set("IT-01") # default value
w2 = OptionMenu(f3,section, "1", "2").grid(column=1,row=5,pady="4")

btn=Button(f3,text="OK",bg="green",fg="white",width="10",font="lucida 10 bold",command=update)
btn.grid(columnspan=3,row=7,pady="20")
f3.pack(pady="115")

''' ************************************************************* Code for  Update Ends ********************************************************'''

''' ************************************************************* Code for Delete starts ********************************************************'''

def delete_details():
    value = tmsg.askquestion("Delete student delails", "Are you sure")
    if value == "yes":
        print("deleting...")
        print(denroll.get(),dsem.get(),dyear.get(),dsection.get())
        delete(denroll.get(),dsem.get()[0],dsection.get())
        
f4=Frame(l,pady="25",padx="25")
l0=Label(f4,text="Delete details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
l2=Label(f4,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
denroll=StringVar()
denl=Entry(f4,textvariable=denroll,width="28").grid(column=1,row=2)

l3=Label(f4,text="Year",font="lucida 10 bold").grid(column=0,row=3,pady="4")
dyear = StringVar()
# dyear.set("1st Year") # default value
w = OptionMenu(f4,dyear, "1st Year", "2nd Year", "3rd Year","4th Year").grid(column=1,row=3,pady="4")

l4=Label(f4,text="Sem",font="lucida 10 bold").grid(column=0,row=4,pady="4")
dsem = StringVar()
# dsem.set("1st sem") # default value
w1 = OptionMenu(f4,dsem,"1st sem","2nd sem","3rd sem","4th sem","5th sem","6th sem","7th sem","8th sem").grid(column=1,row=4,pady="4")

l5=Label(f4,text="Section",font="lucida 10 bold").grid(column=0,row=5,pady="4")
dsection = StringVar()
# dsection.set("IT-01") # default value
w2 = OptionMenu(f4,dsection, "1", "2").grid(column=1,row=5,pady="4")

btn=Button(f4,text="Delete",bg="green",fg="white",width="10",font="lucida 10 bold",command=delete_details)
btn.grid(columnspan=3,row=7,pady="20")
f4.pack(pady="115")

''' ************************************************************* Code for Delete Ends ********************************************************'''
''' ************************************************************* Code for Recording attendance starts ********************************************************'''
def detect1():
    insertdate()
    dt.detect()

def update1():
    up.updatef()

def show1():
    file="data/Attendance_xlsx/third_year_5sem_IT2.xlsx"
    os.startfile("data/Attendance_xlsx/third_year_5sem_IT2.xlsx")

l.pack(ipadx="100",fill=BOTH)

l1=Label(image=photo)

# f98=Frame(l1,pady="25",padx="25")
# ld=Label(f98,text="This is Detect Section",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
# b1=Button(f98,text="Detect",bg="green",fg="white",width="10",font="lucida 10 bold",command=detect1)
# b1.grid(columnspan=3,row=1,pady="20")
# b2=Button(f98,text="Update",bg="green",fg="white",width="10",font="lucida 10 bold",command=update1)
# b2.grid(columnspan=3,row=3,pady="20")
# b3=Button(f98,text="Show",bg="green",fg="white",width="10",font="lucida 10 bold",command=show1)
# b3.grid(columnspan=3,row=5,pady="20")
# f98.pack(pady="120")


f98=Frame(l,pady="25",padx="25")   
l0=Label(f98,text="Enter details to take attendance",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")   

l3=Label(f98,text="SEM : ",font="lucida 10 bold").grid(column=0,row=3,pady="4")

btn=Button(f98,text="Back",bg="green",fg="white",width="10",font="lucida 10 bold",command=back)
btn.grid(columnspan=3,row=7,pady="20")

f98.pack(pady="120")

def insertdate():
    flag=0
    print('Checking if the date is working or not..')
    for i in D.filterdates():
        if str(i.day) == str(datetime.datetime.today().day) and str(i.month) == str(datetime.datetime.today().month) and str(i.year) == str(datetime.datetime.today().year):
            flag=1
    if flag==1:
        wb = load_workbook('data/Attendance_xlsx/third_year_5sem_IT2.xlsx')
        print('Date:',str(i)[:11],' is written in excel and is a working day')
        sheet = wb.active
        current_row = sheet.max_row 
        current_column = sheet.max_column
        print(current_column)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Enrollment"


        current_row = sheet.max_row
        current_column = sheet.max_column
        #sheet.cell(row=1,column=current_column).width = 20
        sheet.cell(row=1, column=current_column+1).value = "".join(str(datetime.datetime.today())[:11])
        
        # save the file 
        wb.save('data/Attendance_xlsx/third_year_5sem_IT2.xlsx') 
    
    else:
        print("this is a holiday popup..ask if they want to continue..")
    



''' ************************************************************* Code for Recording attendance Ends ********************************************************'''

''' ************************************************************* Code for viewing Excel starts ********************************************************'''
def open_excel():
    if(year.get()=="noyear" or sect.get()=="nosec"):
        pass
    else:
        file="data/excel1\\"+year.get()+"_"+sem.get()+"_"+sect.get()+".xlsx"
        os.startfile(file)
        
fa=Frame(l1,pady="8",padx="20" ,height=200)
Label(fa, text = "Select Year",bg="orange",fg="blue",font="lucida 10 bold",width="30").grid(columnspan=3,row=0,pady="10")
year = StringVar()
year.set("first_year")
w1 = OptionMenu(fa,year,"first_year","second_year","third_year","fourth_year").grid(columnspan=3,row=1,pady="4")

Label(fa, text = "Select Semester",bg="orange",fg="blue",font="lucida 10 bold",width="30").grid(columnspan=3,row=2,pady="10")
sem = StringVar()
sem.set("1sem") # default value
w1 = OptionMenu(fa,sem,"1sem","2sem","3sem","4sem","5sem","6sem","7sem","8sem").grid(columnspan=3,row=3,pady="4")


sect = StringVar()
sect.set("nosec")
Label(fa, text = "Select Section",bg="orange",fg="blue",font="lucida 10 bold",width="30").grid(columnspan=3,row=5,pady="10")
radio = Radiobutton(fa, text="IT-1",variable=sect, value="IT1").grid(column=0,row=6,pady="4")
radio = Radiobutton(fa, text="IT-2", padx=14, variable=sect, value="IT2").grid(column=1,row=6,pady="4")
btn=Button(fa,text="show",bg="green",fg="white",width="10",font="lucida 10 bold",command=open_excel)
btn.grid(columnspan=3,row=7,pady="0")
fa.pack(pady="135")
l1.pack(ipadx="150",fill=BOTH)



''' ************************************************************* Code for viewing Excel Ends ********************************************************'''

''' ************************************************************* Code for more details starts ********************************************************'''

l2=Label(image=photo)

f=Frame(l2,pady="25",padx="25")
lbl=Label(f,text="we will help you",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=0)
lbl=Label(f,text="you can contact us on following",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=1)
lbl=Label(f,text="Email  :  unknown@gmail.com",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=2)
lbl=Label(f,text="mobile : 1800-6512-154",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=3)
f.pack(pady="195")

l2.pack(ipadx="100",fill=BOTH)

''' ************************************************************* Code for more details Ends ********************************************************'''

   
root.mainloop()
