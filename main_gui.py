from tkinter import *
from PIL import Image, ImageTk
import tkinter.messagebox as tmsg
import os
from openpyxl import *
from scripts import update as up
from scripts import detect as dt
from scripts import Dates as D
import datetime

root = Tk()
root.geometry("580x550")
root.maxsize(580,550)
root.minsize(580,550)
root.title("Auto Attendance..")

image = Image.open("GUI/register.jpg")
image=image.resize((580,600), Image.ANTIALIAS)
photo = ImageTk.PhotoImage(image)

Admin_id="admin"
Admin_pass="admin@123"
global login_by

'''final code for login section  is start here     '''    
l_login=Label(image=photo)
def Login():
    global login_by
    lid1=lid.get()
    lpass1=lpass.get()
    print(lid.get(),lpass.get())
    if(lid1==Admin_id and lpass1==Admin_pass):
        login_by="Admin"
        login_allow()
    else:
        login_by=""
        login_allow()

  
f_login=Frame(l_login,pady="25",padx="25")
lb0=Label(f_login,text="Enter Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
lb1=Label(f_login,text="Enter ID: ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
lid=StringVar()
e1=Entry(f_login,textvariable=lid,width="28").grid(column=1,row=2)
lb2=Label(f_login,text="Enter Password: ",font="lucida 10 bold").grid(column=0,row=3,pady="4")
lpass=StringVar()
e2=Entry(f_login,textvariable=lpass,width="28").grid(column=1,row=3)
btn=Button(f_login,text="login",bg="green",fg="white",width="10",font="lucida 10 bold",command=Login)
btn.grid(columnspan=3,row=5,pady="10")
f_login.pack(pady="165")

l_login.pack(ipadx="100",fill=BOTH)
'''final code for login section  is end here     '''

def excel(): 

    # resize the width of columns in 
    # excel spreadsheet 
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Enrollment No"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"







def student(x):
    l2.pack_forget()
    l1.pack_forget()
    l.pack(ipadx="100",fill=BOTH)
    if(x==1):
        f4.pack_forget()
        f3.pack_forget()
        f2.pack_forget()
        f21.pack_forget()
        f31.pack_forget()
        f1.pack(pady="110")
    if(x==2):
        f1.pack_forget()
        f3.pack_forget()
        f4.pack_forget()
        f31.pack_forget()
        f21.pack_forget()
        f2.pack(pady="115")
    
    if(x==3):
        f1.pack_forget()
        f2.pack_forget()
        f21.pack_forget()
        f4.pack_forget()
        f31.pack_forget()
        f3.pack(pady="115")
    if(x==4):
        f4.pack(pady="115")
        f1.pack_forget()
        f3.pack_forget()
        f31.pack_forget()
        f2.pack_forget()
        f21.pack_forget()

def attendance(y):
    l.pack_forget()
    l2.pack_forget()
    l1.pack(ipadx="150",fill=BOTH)
    if(y==1):
        fa.pack_forget()
        fd.pack(pady="120")
    if(y==2):
        fd.pack_forget()
        fa.pack(pady="135")

        
def more(z):
    l.pack_forget()
    l1.pack_forget()
    l2.pack(ipadx="100",fill=BOTH)
    
def admin(x):
    if(x==1):
        print("i will train")
    elif(x==2):
        print("i will open excel of teachers")

    
''' code for menu is start here     '''
def login_allow():
    global login_by
    l_login.pack_forget()
    
    mainmenu = Menu(root)
    
    m1 = Menu(mainmenu, tearoff=0)
    m1.add_command(label="Register", command=lambda: student(1))
    m1.add_command(label="View", command=lambda: student(2))
    m1.add_separator()
    m1.add_command(label="Update", command=lambda: student(3))
    m1.add_command(label="Delete", command=lambda: student(4))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="Student", menu=m1)
    
    m2 = Menu(mainmenu, tearoff=0)
    m2.add_command(label="Detect", command=lambda:attendance(1))
    m2.add_separator()
    m2.add_command(label="View Excel", command=lambda:attendance(2))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="Attendance", menu=m2)
   
    if(login_by=="Admin"):
        m4 = Menu(mainmenu, tearoff=0)
        m4.add_command(label="Train",command=lambda:admin(1))
        m4.add_command(label="Teacher Excel",command=lambda:admin(2))
        root.config(menu=mainmenu)
        mainmenu.add_cascade(label="Admin", menu=m4)
    
    m3 = Menu(mainmenu, tearoff=0)
    m3.add_command(label="Help", command=lambda:more(1))
    m3.add_command(label="About Us", command=lambda:more(1))
    root.config(menu=mainmenu)
    mainmenu.add_cascade(label="More", menu=m3)
    
''' code for menu is end here     '''

l = Label(image=photo)
 
''' code for student registration  is start here     '''   
def update_in_attendace_sheet():
    if semester.get() == '1' or semester.get() == '2':
        year = 'first_year'
    elif semester.get() == '3' or semester.get() == '4':
        year = 'second_year'
    elif semester.get() == '5' or semester.get() == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'
    
    wb = load_workbook(f'../data/excel/{year}/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    sheet = wb.active 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 25

    # write given data to an excel spreadsheet 
    
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Enrollment"
    
    
    
    current_row = sheet.max_row 
    current_column = sheet.max_column 

    sheet.cell(row=current_row + 1, column=1).value = name.get() 
    sheet.cell(row=current_row + 1, column=2).value = enroll.get() 
 
     
    wb.save(f'../data/excel/{year}/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    

def update_in_data_sheet():
    #update_in_attendace_sheet()
    # if (e1.get() == "" and
    #     e2.get() == "" and
    #     e3.get() == "" and
    #     e4.get() == "" and
    #     e5.get() == "" and
    #     e6.get() == ""):
    #     print("empty input") 
    if semester.get() == '1' or semester.get() == '2':
        year = 'first_year'
    elif semester.get() == '3' or semester.get() == '4':
        year = 'second_year'
    elif semester.get() == '5' or semester.get() == '6':
        year = 'third_year'
    else:
        year = 'fourth_year'
    
    #wb = load_workbook(f'../data/excel/{year}/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    wb = load_workbook(f'../data/excel/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    sheet = wb.active 
    sheet.column_dimensions['A'].width = 20
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 10
    sheet.column_dimensions['F'].width = 15
    sheet.column_dimensions['G'].width = 25

    # write given data to an excel spreadsheet 
    # at particular location 
    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Enrollment"
    sheet.cell(row=1, column=3).value = "Course"
    sheet.cell(row=1, column=4).value = "Section"
    sheet.cell(row=1, column=5).value = "Semester"
    sheet.cell(row=1, column=6).value = "Contact No"
    sheet.cell(row=1, column=7).value = "Email id"
    
    
    current_row = sheet.max_row 
    current_column = sheet.max_column 

    sheet.cell(row=current_row + 1, column=1).value = name.get()
    sheet.cell(row=current_row + 1, column=2).value = enroll.get() 
    sheet.cell(row=current_row + 1, column=3).value = course.get() 
    sheet.cell(row=current_row + 1, column=4).value = section.get()
    sheet.cell(row=current_row + 1, column=5).value = semester.get()
    sheet.cell(row=current_row + 1, column=6).value = contact.get() 
    sheet.cell(row=current_row + 1, column=7).value = email.get() 
     
    #wb.save(f'../data/excel/{year}/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    wb.save(f'../data/excel/{year}_{semester.get()}sem_IT{section.get()}.xlsx') 
    clear()

    print("form submitted")


def clear(): 

# clear the content of text entry box 
    name.delete(0, END) 
    enroll.delete(0, END) 
    course.delete(0, END) 
    semester.delete(0, END) 
    section.delete(0, END) 
    contact.delete(0, END) 
    email.delete(0, END) 



f1=Frame(l,pady="5",padx="25")
l0=Label(f1,text="Registration Form",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
l1=Label(f1,text="Name : ",font="lucida 10 bold").grid(column=0,row=1,pady="4")
name=Entry(f1,width="28")
name.grid(column=1,row=1)
l2=Label(f1,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
enroll=Entry(f1,width="28")
enroll.grid(column=1,row=2)
l3=Label(f1,text="Course : ",font="lucida 10 bold").grid(column=0,row=3,pady="4")
course=Entry(f1,width="28")
course.grid(column=1,row=3)

l32=Label(f1,text="Section : ",font="lucida 10 bold").grid(column=0,row=4,pady="4")
section=Entry(f1,width="28")
section.grid(column=1,row=4)

l33=Label(f1,text="Sem : ",font="lucida 10 bold").grid(column=0,row=5,pady="4")
semester=Entry(f1,width="28")
semester.grid(column=1,row=5)

l5=Label(f1,text="Contact No : ",font="lucida 10 bold").grid(column=0,row=6,pady="4")
contact=Entry(f1,width="28")
contact.grid(column=1,row=6)

l6=Label(f1,text="Email : ",font="lucida 10 bold").grid(column=0,row=7,pady="4")
email=Entry(f1,width="28")
email.grid(column=1,row=7)
btn=Button(f1,text="Submit",bg="green",fg="white",width="10",font="lucida 10 bold",command=update_in_data_sheet)
btn.grid(columnspan=3,row=8,pady="10")
f1.pack(pady="110")

''' code for student registration  is end here     '''
    
''' code for view student details  is start here     '''

def view():
    
    print(venroll.get(),vsem.get(),vyear.get(),vsection.get())
    f2.pack_forget()
    f21.pack(pady="100")
def back():
    f21.pack_forget()
    f2.pack(pady="115")
    
f21=Frame(l,pady="25",padx="25")   
l0=Label(f21,text="Student Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")   
l1=Label(f21,text="Name : ",font="lucida 10 bold").grid(column=0,row=1,pady="4")
l11=Label(f21,text="from excel",width="28").grid(column=1,row=1)
l2=Label(f21,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
l22=Label(f21,text="from excel",width="28").grid(column=1,row=2)
l3=Label(f21,text="Course : ",font="lucida 10 bold").grid(column=0,row=3,pady="4")
l33=Label(f21,text="from excel",width="28").grid(column=1,row=3)
l4=Label(f21,text="Class : ",font="lucida 10 bold").grid(column=0,row=4,pady="4")
l44=Label(f21,text="from excel",width="28").grid(column=1,row=4)
l5=Label(f21,text="Contact No : ",font="lucida 10 bold").grid(column=0,row=5,pady="4")
l55=Label(f21,text="from excel",width="28").grid(column=1,row=5)
l6=Label(f21,text="Email : ",font="lucida 10 bold").grid(column=0,row=6,pady="4")
l66=Label(f21,text="from excel",width="28").grid(column=1,row=6)
btn=Button(f21,text="Back",bg="green",fg="white",width="10",font="lucida 10 bold",command=back)
btn.grid(columnspan=3,row=7,pady="20")
f21.pack(pady="100")   


f2=Frame(l,pady="25",padx="25")
l0=Label(f2,text="Student details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
l2=Label(f2,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
venroll = StringVar()
e2=Entry(f2,textvariable=venroll,width="28").grid(column=1,row=2)

l3=Label(f2,text="Year",font="lucida 10 bold").grid(column=0,row=3,pady="4")
vyear = StringVar()
vyear.set("1st Year") # default value
w = OptionMenu(f2,vyear, "1st Year", "2nd Year", "3rd Year","4th Year").grid(column=1,row=3,pady="4")

l4=Label(f2,text="Sem",font="lucida 10 bold").grid(column=0,row=4,pady="4")
vsem = StringVar()
vsem.set("1st_sem") # default value
w1 = OptionMenu(f2,vsem,"1st_sem","2nd_sem","3rd_sem","4th_sem","5th_sem","6th_sem","7th_sem","8th_sem").grid(column=1,row=4,pady="4")

l5=Label(f2,text="Section",font="lucida 10 bold").grid(column=0,row=5,pady="4")
vsection = StringVar()
vsection.set("IT-01") # default value
w2 = OptionMenu(f2,vsection, "IT-01", "IT-02").grid(column=1,row=5,pady="4")

btn=Button(f2,text="OK",bg="green",fg="white",width="10",font="lucida 10 bold",command=view)
btn.grid(columnspan=3,row=7,pady="20")
f2.pack(pady="115")


''' code for view student details  is end here     '''

''' code for update student details  is start here     '''


def update():
    f3.pack_forget()
    f31.pack(pady="100")
    
def cancel():
    f31.pack_forget()
    f3.pack(pady="115")
def update_details():
    print(uname.get(),ucourse.get(),ucls.get(),uenrollment.get(),ucontact.get(),uemail.get())
    print("details updated")
    

f31=Frame(l,pady="25",padx="25")
l0=Label(f31,text="Update Details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15") 
l1=Label(f31,text="Name : ",font="lucida 10 bold").grid(column=0,row=1,pady="4")

uname=StringVar()
e1=Entry(f31,textvariable=uname,width="28").grid(column=1,row=1) 
l2=Label(f31,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")

uenrollment=StringVar()
e2=Entry(f31,textvariable=uenrollment,width="28").grid(column=1,row=2) 
l3=Label(f31,text="Course : ",font="lucida 10 bold").grid(column=0,row=3,pady="4")

ucourse=StringVar()   
e3=Entry(f31,textvariable=ucourse,width="28").grid(column=1,row=3) 
l4=Label(f31,text="Class : ",font="lucida 10 bold").grid(column=0,row=4,pady="4")

ucls=StringVar()
e4=Entry(f31,textvariable=ucls,width="28").grid(column=1,row=4) 
l5=Label(f31,text="Contact No : ",font="lucida 10 bold").grid(column=0,row=5,pady="4")

ucontact=StringVar()
e5=Entry(f31,textvariable=ucontact,width="28").grid(column=1,row=5) 
l6=Label(f31,text="Email : ",font="lucida 10 bold").grid(column=0,row=6,pady="4")

uemail=StringVar()
e6=Entry(f31,textvariable=uemail,width="28").grid(column=1,row=6) 
btn=Button(f31,text="Submit",bg="green",fg="white",width="10",font="lucida 10 bold",command=update_details)
btn.grid(column=0,row=7,pady="20")    
btn1=Button(f31,text="cancel",bg="green",fg="white",width="10",font="lucida 10 bold",command=cancel)
btn1.grid(column=1,row=7,pady="20")    
f31.pack(pady="100")


f3=Frame(l,pady="25",padx="25")
l0=Label(f3,text="Update details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
l2=Label(f3,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
e2=Entry(f3,width="28").grid(column=1,row=2)

l3=Label(f3,text="Year",font="lucida 10 bold").grid(column=0,row=3,pady="4")
year = StringVar()
year.set("1st Year") # default value
w = OptionMenu(f3,year, "1st Year", "2nd Year", "3rd Year","4th Year").grid(column=1,row=3,pady="4")

l4=Label(f3,text="Sem",font="lucida 10 bold").grid(column=0,row=4,pady="4")
sem = StringVar()
sem.set("1st sem") # default value
w1 = OptionMenu(f3,sem,"1st sem","2nd sem","3rd sem","4th sem","5th sem","6th sem","7th sem","8th sem").grid(column=1,row=4,pady="4")

l5=Label(f3,text="Section",font="lucida 10 bold").grid(column=0,row=5,pady="4")
section = StringVar()
section.set("IT-01") # default value
w2 = OptionMenu(f3,section, "IT-01", "IT-02").grid(column=1,row=5,pady="4")

btn=Button(f3,text="OK",bg="green",fg="white",width="10",font="lucida 10 bold",command=update)
btn.grid(columnspan=3,row=7,pady="20")
f3.pack(pady="115")

''' code for update student details  is end here     '''

''' code for delete student details  is start here     '''

def delete():
    value = tmsg.askquestion("Delete student delails", "Are you sure")
    if value == "yes":
        print("detail deleted")
        print(denroll.get(),dsem.get(),dyear.get(),dsection.get())
        
f4=Frame(l,pady="25",padx="25")
l0=Label(f4,text="Delete details",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
l2=Label(f4,text="Enrollment No : ",font="lucida 10 bold").grid(column=0,row=2,pady="4")
denroll=StringVar()
denl=Entry(f4,textvariable=denroll,width="28").grid(column=1,row=2)

l3=Label(f4,text="Year",font="lucida 10 bold").grid(column=0,row=3,pady="4")
dyear = StringVar()
dyear.set("1st Year") # default value
w = OptionMenu(f4,dyear, "1st Year", "2nd Year", "3rd Year","4th Year").grid(column=1,row=3,pady="4")

l4=Label(f4,text="Sem",font="lucida 10 bold").grid(column=0,row=4,pady="4")
dsem = StringVar()
dsem.set("1st sem") # default value
w1 = OptionMenu(f4,dsem,"1st sem","2nd sem","3rd sem","4th sem","5th sem","6th sem","7th sem","8th sem").grid(column=1,row=4,pady="4")

l5=Label(f4,text="Section",font="lucida 10 bold").grid(column=0,row=5,pady="4")
dsection = StringVar()
dsection.set("IT-01") # default value
w2 = OptionMenu(f4,dsection, "IT-01", "IT-02").grid(column=1,row=5,pady="4")

btn=Button(f4,text="Delete",bg="green",fg="white",width="10",font="lucida 10 bold",command=delete)
btn.grid(columnspan=3,row=7,pady="20")
f4.pack(pady="115")

''' code for delete student details  is end here     '''

l.pack(ipadx="100",fill=BOTH)

'''final code for student section  is end here     '''
'''final code for attendance section  is start here     '''

l1=Label(image=photo)
'''code for Detect  is start here     '''

def insertdate():
    flag=0
    for i in D.filterdates():
        if str(i.day) == str(datetime.datetime.today().day) and str(i.month) == str(datetime.datetime.today().month) and str(i.year) == str(datetime.datetime.today().year):
            flag=1
    if flag==1:
        wb = load_workbook('./data/Attendance_xlsx/third_year_5sem_IT2.xlsx')
        print('Date:',str(i)[:11],' is written in excel and is a working day')
        sheet = wb.active
        current_row = sheet.max_row 
        current_column = sheet.max_column
        print(current_column)
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.cell(row=1, column=1).value = "Name"
        sheet.cell(row=1, column=2).value = "Enrollment"


        current_row = sheet.max_row
        current_column = sheet.max_column
        #sheet.cell(row=1,column=current_column).width = 20
        sheet.cell(row=1, column=current_column+1).value = "".join(str(datetime.datetime.today())[:11])
        
        # save the file 
        wb.save('./data/Attendance_xlsx/third_year_5sem_IT2.xlsx') 
    
    else:
        print("this is a holiday popup..ask if they want to continue..")



def detect1():
    insertdate()
    dt.detect()


def update1():
    up.updatef()

def show1():
    file="data/Attendance_xlsx/third_year_5sem_IT2.xlsx"
    os.startfile("data/Attendance_xlsx/third_year_5sem_IT2.xlsx")
    

fd=Frame(l1,pady="25",padx="25")
ld=Label(fd,text="This is Detect Section",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(columnspan=3,row=0,pady="15")
b1=Button(fd,text="Detect",bg="green",fg="white",width="10",font="lucida 10 bold",command=detect1)
b1.grid(columnspan=3,row=1,pady="20")
b2=Button(fd,text="Update",bg="green",fg="white",width="10",font="lucida 10 bold",command=update1)
b2.grid(columnspan=3,row=3,pady="20")
b3=Button(fd,text="Show",bg="green",fg="white",width="10",font="lucida 10 bold",command=show1)
b3.grid(columnspan=3,row=5,pady="20")
fd.pack(pady="120")

'''code for Detect  is end here     '''

'''code for view excel  is start here     '''
def open_excel():
    if(year.get()=="noyear" or sect.get()=="nosec"):
        pass
    else:
        file="data/excel1\\"+year.get()+"_"+sem.get()+"_"+sect.get()+".xlsx"
        os.startfile(file)
        
fa=Frame(l1,pady="8",padx="20" ,height=200)
Label(fa, text = "Select Year",bg="orange",fg="blue",font="lucida 10 bold",width="30").grid(columnspan=3,row=0,pady="10")
year = StringVar()
year.set("first_year")
w1 = OptionMenu(fa,year,"first_year","second_year","third_year","fourth_year").grid(columnspan=3,row=1,pady="4")

Label(fa, text = "Select Semester",bg="orange",fg="blue",font="lucida 10 bold",width="30").grid(columnspan=3,row=2,pady="10")
sem = StringVar()
sem.set("1sem") # default value
w1 = OptionMenu(fa,sem,"1sem","2sem","3sem","4sem","5sem","6sem","7sem","8sem").grid(columnspan=3,row=3,pady="4")


sect = StringVar()
sect.set("nosec")
Label(fa, text = "Select Section",bg="orange",fg="blue",font="lucida 10 bold",width="30").grid(columnspan=3,row=5,pady="10")
radio = Radiobutton(fa, text="IT-1",variable=sect, value="IT1").grid(column=0,row=6,pady="4")
radio = Radiobutton(fa, text="IT-2", padx=14, variable=sect, value="IT2").grid(column=1,row=6,pady="4")
btn=Button(fa,text="show",bg="green",fg="white",width="10",font="lucida 10 bold",command=open_excel)
btn.grid(columnspan=3,row=7,pady="0")
fa.pack(pady="135")
'''code for view excel  is end here     '''

l1.pack(ipadx="150",fill=BOTH)

'''final code for attendance section  is end here     '''
'''final code for More section  is start here     '''
l2=Label(image=photo)

f=Frame(l2,pady="25",padx="25")
lbl=Label(f,text="we will help you",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=0)
lbl=Label(f,text="you can contact us on following",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=1)
lbl=Label(f,text="Email  :  unknown@gmail.com",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=2)
lbl=Label(f,text="mobile : 1800-6512-154",bg="orange",fg="blue",font="lucida 10 bold",width="35",pady="4").grid(column=0,row=3)
f.pack(pady="195")

l2.pack(ipadx="100",fill=BOTH)
'''final code for More section  is end here     '''

'''final code for admin section  is start here     '''    


'''final code for admin section  is end here     '''
root.mainloop()
